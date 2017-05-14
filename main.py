
import requests
import xlsxwriter
import re
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from openpyxl import load_workbook


def scrap(url, a):
 try:
  page=requests.get(url,verify=False)
 except ValueError:
  print("Not a valid url")
  exit()

 soup = BeautifulSoup(page.content, 'html.parser')


#getting <a> elements
 text=list(soup.find_all('a',href=True ))
 f=open('scrap.txt','w',encoding='utf8')
 for s in text:
         f.write(s.get_text().strip())
         f.write("\n ")
 f.write("***")
 f.write("\n")



#getting <href> elements
 text=list(soup.find_all('a',href=True))
 for s in text:
          f.write(urljoin(url,s['href']).strip())
          f.write("\n")
 f.write("***")
 f.write("\n")

#getting <h2> elements
 text=list(soup.find_all('h2'))
 for s in text:
         f.write(s.get_text().strip())
         f.write("\n")
 f.write("***")
 f.write("\n")

#getting <li> elements
 text=list(soup.find_all('li'))
 for s in text:
          f.write(s.get_text().strip())
          f.write("\n")
 f.write("***")
 f.write("\n")


#getting <p> elements
 text=list(soup.find_all('p'))
 for s in text:
          f.write(s.get_text().strip())    #strip() is used to remove \n character
          f.write("\n")
 f.close()

 

#opening a xlsx workbook
 string=str(a)
 wb2 = load_workbook('t.xlsx')
 ws1=wb2.create_sheet(string)
 
 #ws1=wb2.get_sheet_by_name('sheet1')
 file=open('currenturl.txt','w')
 file.write(url)
 file.close()
 row=1
 col=1
 a="***" # delimeter for different tags

#writing to the worksheet
 ws1.cell(row=row,column=col).value="<a> elements"
 row=row+1
 f=open('scrap.txt','r',encoding='utf8')


 for s in f :
     if(s.strip()==a) : 
                 row=1
                 col=col+1
                 if(col==1):     
                      ws1.cell(row=row,column=col).value="<href> elements"
                 elif(col==2):
                      ws1.cell(row=row,column=col).value="<h2> elements"
                 elif(col==3):
                      ws1.cell(row=row,column=col).value="<li> elements"
                 else:
                      ws1.cell(row=row,column=col).value="<p> elements"
                 row=row+1
                 continue
     ws1.cell(row=row,column=col).value=s.strip()
     row=row+1
 wb2.save('t.xlsx')
 wb2.close()

 f.close()
